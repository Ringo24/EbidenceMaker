import os
import pyautogui
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import rows_from_range
import tkinter as tk
from tkinter import filedialog

class CaptureApp:
    def __init__(self, master):
        self.master = master
        master.title("화면 캡처 프로그램")

        # 엑셀 파일 경로 입력 텍스트 상자
        self.excel_path_label = tk.Label(master, text="엑셀 파일 경로:")
        self.excel_path_label.grid(row=0, column=0, sticky='w')

        self.excel_path_entry = tk.Entry(master, width=50)
        self.excel_path_entry.grid(row=0, column=1)

        # 파일 불러오기 버튼
        self.load_excel_button = tk.Button(master, text="파일 불러오기", command=self.load_excel)
        self.load_excel_button.grid(row=0, column=2)

        # 단축키 입력 텍스트 상자
        self.hotkey_label = tk.Label(master, text="단축키:")
        self.hotkey_label.grid(row=1, column=0, sticky='w')

        self.hotkey_entry = tk.Entry(master)
        self.hotkey_entry.grid(row=1, column=1)

        # 이미지 간의 행 간격 입력 텍스트 상자
        self.row_interval_label = tk.Label(master, text="이미지 간의 행 간격:")
        self.row_interval_label.grid(row=2, column=0, sticky='w')

        self.row_interval_entry = tk.Entry(master)
        self.row_interval_entry.grid(row=2, column=1)

        # 트레이 모드로 전환 버튼
        self.tray_mode_button = tk.Button(master, text="트레이 모드로 전환", command=self.tray_mode)
        self.tray_mode_button.grid(row=3, column=0, columnspan=2)

        # 종료 버튼
        self.quit_button = tk.Button(master, text="종료", command=master.quit)
        self.quit_button.grid(row=3, column=2)

    def load_excel(self):
        excel_filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        self.excel_path_entry.delete(0, tk.END)
        self.excel_path_entry.insert(0, excel_filepath)

    def capture_active_window(self):
        active_window = gw.getActiveWindow()
        left, top, width, height = active_window.left, active_window.top, active_window.width, active_window.height
        screenshot = pyautogui.screenshot(region=(left, top, width, height))
        return screenshot

    def insert_image(self, excel_filepath, screenshot, row_interval):
        if os.path.exists(excel_filepath):
            wb = load_workbook(excel_filepath)
        else:
            wb = Workbook()
            wb.save(excel_filepath)

        ws = wb.active
        row = 1

        img_path = 'screenshot.png'
        screenshot.save(img_path)

        img = Image(img_path)
        for r in rows_from_range(ws.dimensions):
            if row <= r[0]:
                ws.add_image(img, f'A{r[0]}')
                row = r[0] + int(row_interval)
                break

        wb.save(excel_filepath)

    def capture_and_insert_screenshot(self):
        excel_filepath = self.excel_path_entry.get()
        hotkey = tuple(self.hotkey_entry.get().split(','))
        row_interval = int(self.row_interval_entry.get())

        while True:
            screenshot = self.capture_active_window()
            self.insert_image(excel_filepath, screenshot, row_interval)

    def tray_mode(self):
        # 트레이 모드로 전환하는 함수 작성
        pass

root = tk.Tk()
app = CaptureApp(root)
root.mainloop()
